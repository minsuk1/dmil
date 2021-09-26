# 카테고리 dictionary형태
category={"네일케어": 50000196, "클렌징": 50000192, 
         "색조메이크업": 50000195, "바디케어":50000197, 
         "뷰티소품": 50000201, "베이스메이크업": 50000194, 
         "헤어스타일링":50000199, "향수":50000200, 
         "스킨/바디용품": 50000002, "스킨케어":50000190, 
         "남성화장품":50000202, "헤어케어":50000198, 
         "선케어":50000191, "마스크/팩":50000193}

#----------------------------------------------------------------------

# 엑셀에서 데이터 가져와서 이중리스트로 만들기
from openpyxl import load_workbook

# test.xlsx데이터 위치(경로)
dir = "C:/Users/palt1/Desktop/portfolio/test.xlsx"

excel = load_workbook(dir)
sheet = excel['sheet1']

keywords=[]
categories=[]
categories_catId=[]
        
# 키워드 불러오기
for cell in sheet['A']:
    keywords.append(cell.value)

# 카테고리 불러오기
for cell in sheet['B']:
    categories.append(cell.value)
    
# 카테고리 -> catId 변경
for i in categories:
    categories_catId.append(category[i])
    
# 키워드 + 카테고리
total=[]
for i in zip(keywords,categories_catId):
    total.append(i)


import requests
import pandas as pd
from bs4 import BeautifulSoup

for i in total:
    tmp=sol(i[0],i[1])
    dataFrame = pd.DataFrame(data = tmp, columns = ['제목'])
    dataFrame.to_excel('C:/Users/palt1/Desktop/portfolio/{0}.xlsx'.format(i[0]), index=False, sheet_name=i[0])


#----------------------------------------------------------------------

# 크롤링 후 해당 데이터를 excel로 저장

import requests
from bs4 import BeautifulSoup

def sol(keyword,category):
    total_data=[]
    for page_num in range(4):
        url=f"https://search.shopping.naver.com/search/all?&frm=NVSHCAT&pagingSize=40&productSet=total&sort=rel&timestamp=&viewType=list"
        # 스킨/바디용품일 때는 query

        payload={"pagingindex":page_num, "catId":category, "origQuery":keyword, "query":keyword,}
        res = requests.get(url,params=payload)

        soup = BeautifulSoup(res.content, 'html.parser')

        data = soup.select('div > a.basicList_link__1MaTN')
        for item in data:
            total_data.append(item.get_text().strip())
            
    return total_data


import requests
import pandas as pd
from bs4 import BeautifulSoup

for i in total:
    tmp=sol(i[0],i[1])
    dataFrame = pd.DataFrame(data = tmp, columns = ['제목'])
    dataFrame.to_excel('C:/Users/palt1/Desktop/portfolio/{0}.xlsx'.format(i[0]), index=False, sheet_name=i[0])


#----------------------------------------------------------------------

# 엑셀에 있는 데이터를 pdf로 변환

def excelInfo(filepath): 
    excel_list = os.listdir(filepath) 
    result = [] 
    for file in excel_list: 
        wb = op.load_workbook(filepath+"/"+file)
        ws_list = wb.sheetnames 
        filename = file.replace(".xlsx","") 
        
        for sht in ws_list: 
            temp_tuple = (filepath+"/"+file, filename, sht)
            result.append(temp_tuple) 
    print(result) 
    return result 


def transPDF(fileinfo, savepath): 
    excel = win32com.client.Dispatch("Excel.Application") 
    i=0 
    for info in fileinfo: 
        wb = excel.Workbooks.Open(info[0]) 
        ws = wb.Worksheets(info[2]) 
        ws.Select()
        
        wb.ActiveSheet.ExportAsFixedFormat(0, savepath+"/"+str(i)+"_"+info[1]+"_"+info[2]+".pdf")
        i=i+1 
        wb.Close(False) 
        excel.Quit()


import win32com.client 
import openpyxl as op 
import os


filepath = "C:/Users/palt1/Desktop/tmp" 
pdfpath = "C:/Users/palt1/Desktop/tmp" 
excelinfo = excelInfo(filepath) 
transPDF(excelinfo, pdfpath)

